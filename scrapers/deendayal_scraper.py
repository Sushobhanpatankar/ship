"""
Deendayal Port Authority (formerly Kandla Port) — vessel berth scraper.

Data source: Daily Berthing List Excel file published each morning at:
  https://www.deendayalport.gov.in/wp-content/uploads/YYYY/MM/Daily-Berthing-List-DD.MM.YYYY.xlsx

The workbook has five sheets:
  AT BERTH   — currently berthed vessels (A1:R49) ← primary
  WAITING    — vessels at anchorage waiting for berth
  ETA        — expected arrivals

Column layout for AT BERTH / WAITING / ETA (row 6 = header):
  A: SR NO  B: PRIORITY  C: BERTH  D: VCN NO.  E: VESSEL NAME
  F: PANEL  G: TIDE      H: LOA    I: I/E       J: CARGO
  K: QTY    L: UOM       M: NORMS  N: MADE FAST O: COMM
  P: ETC    Q: AGENT     R: REMARKS

CARGO column (J) contains explicit commodity text, e.g.:
  CRUDE OIL, CRUDE, LPG, LNG, HSD, PETROLEUM, NAPHTHA, FURNACE OIL,
  PETCOKE, COAL, SALT, AGGREGATES, TIMBER …

Energy tanker berths at Kandla/Deendayal:
  SPM-1, SPM-2  — crude oil VLCC/Suezmax (Single Point Mooring, offshore)
  IOC, HPCL, BPCL Jetty  — petroleum products
  L-1 … L-6    — LPG jetties
"""
import asyncio
import datetime
import io
import logging

import httpx
import openpyxl

from scrapers.base_scraper import BaseScraper

log = logging.getLogger(__name__)

_BASE_URL = "https://www.deendayalport.gov.in/wp-content/uploads/{year}/{month:02d}/Daily-Berthing-List-{day:02d}.{month:02d}.{year}.xlsx"
_LISTING_URL = "https://www.deendayalport.gov.in/en/berthing_status/"

# Berth prefix → cargo override
_BERTH_CARGO_MAP = {
    "SPM":  "CRUDE",
    "IOC":  "PETROLEUM",
    "HPCL": "PETROLEUM",
    "BPCL": "PETROLEUM",
    "L-":   "PETROLEUM",  # LPG jetties
}

# Sheets to parse (in priority order)
_SHEETS = ["AT BERTH", "WAITING", "ETA"]


class DeendayalScraper(BaseScraper):
    name = "Deendayal"
    url = _LISTING_URL
    port_name = "Deendayal"

    # Override run() — we download xlsx directly, not HTML
    async def run(self) -> list[dict]:
        try:
            xlsx_bytes = await self._fetch_xlsx()
            records = _parse_xlsx(xlsx_bytes, self)
            from database import db
            await db.log_scraper_run(self.name, "OK", len(records))
            log.info("[Deendayal] scraped %d energy vessel records", len(records))
            return records
        except Exception as e:
            try:
                from database import db
                await db.log_scraper_run(self.name, "ERROR", 0, str(e))
            except Exception:
                pass
            log.warning("[Deendayal] scrape failed: %s", e)
            return []

    # parse() required by ABC — not used when run() is overridden
    async def parse(self, html: str) -> list[dict]:  # pragma: no cover
        return []

    async def _fetch_xlsx(self) -> bytes:
        """Try today then up to 2 previous days (file may not be uploaded yet)."""
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
        }
        today = datetime.date.today()
        async with httpx.AsyncClient(timeout=30, follow_redirects=True, headers=headers) as client:
            for delta in range(3):  # try today, yesterday, day before
                d = today - datetime.timedelta(days=delta)
                url = _BASE_URL.format(year=d.year, month=d.month, day=d.day)
                try:
                    resp = await client.get(url)
                    if resp.status_code == 200 and resp.content[:4] in (b"PK\x03\x04", b"\xd0\xcf\x11\xe0"):
                        log.info("[Deendayal] downloaded xlsx for %s (%d bytes)", d, len(resp.content))
                        return resp.content
                    log.debug("[Deendayal] xlsx not found for %s (status=%d)", d, resp.status_code)
                except httpx.HTTPError as e:
                    log.debug("[Deendayal] HTTP error for %s: %s", d, e)

        raise RuntimeError("Deendayal: no xlsx found for past 3 days")


# ─────────────────────────────────────────────────────────────
# Excel parser
# ─────────────────────────────────────────────────────────────

def _parse_xlsx(xlsx_bytes: bytes, scraper: BaseScraper) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    records: list[dict] = []
    seen: set[str] = set()

    for sheet_name in _SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find the header row (contains VESSEL NAME)
        hdr_idx = None
        hdr = []
        for i, row in enumerate(rows):
            vals = [str(x).upper() if x else "" for x in row]
            if any("VESSEL" in v for v in vals):
                hdr_idx = i
                hdr = vals
                break

        if hdr_idx is None:
            continue

        # Locate key columns by header content
        col_berth  = _find_col(hdr, ["BERTH"])
        col_vessel = _find_col(hdr, ["VESSEL NAME", "VESSEL"])
        col_ie     = _find_col(hdr, ["I/E"])
        col_cargo  = _find_col(hdr, ["CARGO"])
        col_etc    = _find_col(hdr, ["ETC", "ETA"])
        col_agent  = _find_col(hdr, ["AGENT"])
        col_qty    = _find_col(hdr, ["QTY", "QUANTITY"])
        col_uom    = _find_col(hdr, ["UOM", "UNIT"])

        if col_vessel is None or col_cargo is None:
            continue

        for row in rows[hdr_idx + 1:]:
            if not row or not row[col_vessel]:
                continue

            vessel = str(row[col_vessel]).strip()
            if not vessel or len(vessel) < 3 or vessel.startswith("="):
                continue
            if vessel.upper() in ("-", "NIL", "N/A", "TBA", "VACANT"):
                continue

            # Deduplicate by vessel name across sheets
            key = vessel.upper()
            if key in seen:
                continue
            seen.add(key)

            berth      = str(row[col_berth]).strip() if col_berth is not None and row[col_berth] else ""
            cargo_text = str(row[col_cargo]).strip() if row[col_cargo] else ""
            ie_text    = str(row[col_ie]).strip() if col_ie is not None and row[col_ie] else ""
            etc_val    = row[col_etc] if col_etc is not None else None
            etc_str    = _fmt_date(etc_val)
            raw_qty    = row[col_qty] if col_qty is not None and row[col_qty] else None
            uom_str    = str(row[col_uom]).strip().upper() if col_uom is not None and row[col_uom] else "MT"

            # Classify cargo
            cargo_cat = scraper.detect_cargo(cargo_text + " " + vessel)
            if cargo_cat == "OTHER":
                cargo_cat = _berth_to_cargo(berth)
            if cargo_cat == "OTHER":
                continue  # skip non-energy vessels

            activity = "BERTHED" if sheet_name == "AT BERTH" else "ANCHORED"

            records.append({
                "ship_name":          vessel,
                "berth":              berth,
                "cargo_category":     cargo_cat,
                "activity":           activity,
                "arrival_time":       "",
                "expected_departure": etc_str,
                "quantity_mt":        _qty_to_mt(raw_qty, uom_str),
                "port_name":          scraper.port_name,
                "source":             scraper.name,
            })

    log.debug("[Deendayal] parsed %d energy vessels across sheets", len(records))
    return records


def _find_col(headers: list[str], keywords: list[str]) -> int | None:
    for i, h in enumerate(headers):
        if any(kw in h for kw in keywords):
            return i
    return None


def _berth_to_cargo(berth: str) -> str:
    b = berth.upper().strip()
    for prefix, cat in _BERTH_CARGO_MAP.items():
        if b.startswith(prefix):
            return cat
    return "OTHER"


def _qty_to_mt(raw, uom: str) -> int:
    """Convert a raw quantity value and its UOM to metric tonnes (integer)."""
    if raw is None:
        return 0
    try:
        val = float(str(raw).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0
    uom = uom.upper().strip()
    if uom in ("MT", "MTS", "M.T.", "METRIC TON", "METRIC TONNE"):
        return int(val)
    if uom in ("KL", "KLS", "KILOLITRE", "KILOLITER"):
        return int(val * 0.86)  # approx density for crude/petroleum
    if uom in ("KGS", "KG", "KILOGRAM", "KILOGRAMS"):
        return int(val / 1000)
    return 0  # unknown UOM — safe default


def _fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d-%m-%Y %H:%M") if isinstance(val, datetime.datetime) else val.strftime("%d-%m-%Y")
    return str(val).strip()
